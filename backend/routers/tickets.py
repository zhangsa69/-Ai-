"""Ticket endpoints: list, detail, resend, export, image upload."""
import io
import logging
import uuid
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
)
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from auth import get_current_user
from config import settings
from database import get_db
from models import AdminUser, SystemConfig, Ticket
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from schemas import (
    ImageUploadResponse,
    TicketListResponse,
    TicketOut,
    TicketResendResponse,
)
from services.email_service import send_ticket_email

log = logging.getLogger("tickets")
router = APIRouter()

UPLOAD_DIR = Path(__file__).resolve().parent.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}


def _to_out(t: Ticket) -> TicketOut:
    return TicketOut(
        id=t.id,
        conversation_id=t.conversation_id,
        name=t.name,
        phone=t.phone,
        location=t.location,
        description=t.description,
        images=[p for p in (t.images or "").split(",") if p.strip()],
        category=t.category,
        email_status=t.email_status,
        email_to=t.email_to,
        email_error=t.email_error or "",
        resend_count=t.resend_count,
        created_at=t.created_at,
        updated_at=t.updated_at,
    )


@router.get("", response_model=TicketListResponse, summary="获取工单列表", description="""
分页查询所有工单，可按故障分类或邮件状态筛选。

**筛选参数（都是可选的，不传则查全部）：**
- `category` — 故障分类：`desktop`(桌面运维) / `network`(网络) / `weak_current`(弱电) / `unknown`(未分类)
- `email_status` — 邮件状态：`pending`(待发送) / `sent`(已发送) / `failed`(发送失败)
- `page` — 页码（从第 1 页开始，默认 1）
- `page_size` — 每页条数（默认 20，最多 200）

> 🔒 需要管理员登录
""")
async def list_tickets(
    category: Optional[str] = Query(None, description="按分类筛选: desktop/network/weak_current/unknown"),
    email_status: Optional[str] = Query(None, description="按邮件状态筛选: pending/sent/failed"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    _user: AdminUser = Depends(get_current_user),
):
    """List tickets (admin only) with optional category & status filters and pagination."""
    stmt = select(Ticket)
    count_stmt = select(func.count(Ticket.id))

    if category:
        stmt = stmt.where(Ticket.category == category)
        count_stmt = count_stmt.where(Ticket.category == category)
    if email_status:
        stmt = stmt.where(Ticket.email_status == email_status)
        count_stmt = count_stmt.where(Ticket.email_status == email_status)

    total = (await db.execute(count_stmt)).scalar() or 0
    stmt = stmt.order_by(Ticket.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(stmt)).scalars().all()

    return TicketListResponse(total=total, items=[_to_out(t) for t in rows])


@router.get("/{ticket_id}", response_model=TicketOut, summary="查看单个工单详情", description="""
根据工单 ID 查看详细信息，包括故障描述、分类、邮件状态、图片等。

> 🔒 需要管理员登录
""")
async def get_ticket(
    ticket_id: int,
    db: AsyncSession = Depends(get_db),
    _user: AdminUser = Depends(get_current_user),
):
    result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="工单不存在")
    return _to_out(t)


@router.post("/{ticket_id}/resend", response_model=TicketResendResponse, summary="重新发送工单邮件", description="""
手动重发某个工单的通知邮件。通常用于上次发送失败后重试。

重发次数会自动 +1，邮件状态会更新为 `sent`（成功）或 `failed`（失败）。

> 🔒 需要管理员登录
""")
async def resend_ticket(
    ticket_id: int,
    db: AsyncSession = Depends(get_db),
    _user: AdminUser = Depends(get_current_user),
):
    """Manually resend the ticket notification email."""
    result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="工单不存在")

    cfg_map = {
        cfg.key: cfg.value
        for cfg in (await db.execute(select(SystemConfig))).scalars().all()
    }
    success, err = await send_ticket_email(t, cfg_map)
    t.email_status = "sent" if success else "failed"
    t.email_error = err or ""
    t.resend_count = (t.resend_count or 0) + 1
    await db.commit()
    return TicketResendResponse(
        success=success,
        message="发送成功" if success else f"发送失败: {err}",
        email_status=t.email_status,
    )


@router.get("/export/all.xlsx", summary="导出工单为 Excel", description="""
将所有工单导出为 .xlsx 格式的 Excel 文件，可直接用 WPS 或 Excel 打开。

导出的列包括：ID、姓名、手机号、故障位置、故障描述、故障分类、收件邮箱、邮件状态、重发次数、创建时间、更新时间。

> 🔒 需要管理员登录
""")
async def export_excel(
    db: AsyncSession = Depends(get_db),
    _user: AdminUser = Depends(get_current_user),
):
    """Export all tickets as an .xlsx file."""
    result = await db.execute(select(Ticket).order_by(Ticket.created_at.desc()))
    tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "工单列表"

    headers = [
        "ID", "姓名", "手机号", "故障位置", "故障描述", "故障分类",
        "收件邮箱", "邮件状态", "重发次数", "创建时间", "更新时间",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    category_label = {"desktop": "桌面运维", "network": "网络", "weak_current": "弱电", "unknown": "未分类"}
    status_label = {"pending": "待发送", "sent": "已发送", "failed": "失败"}

    for t in tickets:
        ws.append([
            t.id,
            t.name,
            t.phone,
            t.location,
            t.description,
            category_label.get(t.category, t.category),
            t.email_to,
            status_label.get(t.email_status, t.email_status),
            t.resend_count,
            t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
            t.updated_at.strftime("%Y-%m-%d %H:%M:%S") if t.updated_at else "",
        ])

    # Column widths
    widths = [6, 12, 16, 20, 40, 12, 24, 12, 10, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/upload", response_model=ImageUploadResponse)
async def upload_image(
    file: UploadFile = File(...),
    conversation_id: Optional[str] = None,
    _user: AdminUser = Depends(get_current_user),
):
    """Upload a fault image (admin). Note: this endpoint requires auth.

    For user-facing uploads during the chat flow, we use a separate endpoint
    `POST /api/tickets/public-upload` (no auth required) below.
    """
    raise HTTPException(status_code=410, detail="Use /api/tickets/public-upload instead")


@router.post("/public-upload", response_model=ImageUploadResponse, summary="上传故障图片（公开）", description="""
用户在对话页面上传故障现场截图。无需登录即可使用。

**限制：**
- 支持格式：jpg / jpeg / png / gif / webp / bmp
- 最大文件大小：10MB
- 上传后返回可访问的图片 URL
""")
async def public_upload_image(
    file: UploadFile = File(...),
    conversation_id: Optional[str] = None,
):
    """Public image upload endpoint used by the chat page (no auth required)."""
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型: {ext}")

    # Cap file size
    max_bytes = settings.max_upload_size_mb * 1024 * 1024
    contents = await file.read()
    if len(contents) > max_bytes:
        raise HTTPException(status_code=413, detail=f"文件超过 {settings.max_upload_size_mb}MB 限制")

    # Sub-directory by conversation_id
    sub = (conversation_id or "misc").replace("/", "_").replace("\\", "_")
    target_dir = UPLOAD_DIR / sub
    target_dir.mkdir(parents=True, exist_ok=True)

    name = f"{uuid.uuid4().hex[:12]}{ext}"
    target = target_dir / name
    with open(target, "wb") as f:
        f.write(contents)

    relative = f"{sub}/{name}"
    return ImageUploadResponse(
        url=f"/uploads/{relative}",
        filename=relative,
    )
